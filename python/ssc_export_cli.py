#!/usr/bin/env python3
import requests
import pandas as pd
import re
import base64
import argparse
import urllib3
import logging
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ============================================================
# Helper functions
# ============================================================

def clean_html(raw_text):
    """Hilangkan tag HTML dari teks"""
    return re.sub(r"<.*?>", "", raw_text).strip() if raw_text else ""

def severity_rank(sev):
    """Urutkan severity: Critical > High > Medium > Low > Info"""
    order = {"critical": 4, "high": 3, "medium": 2, "low": 1, "info": 0}
    return order.get(str(sev).strip().lower(), -1)

def severity_color(sev):
    """Warna baris berdasarkan severity"""
    sev = str(sev).lower()
    if sev == "critical":
        return "FFC7CE"  # merah muda
    elif sev == "high":
        return "FFEB9C"  # kuning-oranye
    elif sev == "medium":
        return "FFF2CC"  # kuning muda
    elif sev == "low":
        return "C6EFCE"  # hijau muda
    elif sev == "info":
        return "BDD7EE"  # biru muda
    return None

# ============================================================
# Fortify API functions
# ============================================================

def authenticate(base_url, username, password, verify_tls, debug=False):
    """Minta UnifiedLoginToken dari Fortify SSC"""
    url = f"{base_url}/api/v1/tokens"
    creds = f"{username}:{password}".encode("utf-8")
    basic_auth = base64.b64encode(creds).decode("utf-8")

    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Authorization": f"Basic {basic_auth}"
    }
    payload = {"type": "UnifiedLoginToken", "description": "API Excel Export"}

    resp = requests.post(url, headers=headers, json=payload, verify=verify_tls)
    if resp.status_code != 201:
        raise SystemExit(f"❌ Auth gagal ({resp.status_code}): {resp.text[:200]}")

    token = resp.json()["data"]["token"]
    if debug:
        print("✅ Token berhasil diperoleh:", token[:20], "...")
    return {"token_type": "FortifyToken", "access_token": token}


def fetch_issues(base_url, project_version, auth, verify_tls, debug=False):
    """Ambil semua issue (tanpa batasan engine type) dengan pagination"""
    headers = {
        "Authorization": f"{auth['token_type']} {auth['access_token']}",
        "Accept": "application/json"
    }

    all_issues = []
    start = 0
    limit = 500

    while True:
        url = f"{base_url}/api/v1/projectVersions/{project_version}/issues?start={start}&limit={limit}"
        resp = requests.get(url, headers=headers, verify=verify_tls)
        if resp.status_code != 200:
            raise SystemExit(f"❌ Gagal ambil issues ({resp.status_code}): {resp.text[:200]}")

        data = resp.json().get("data", [])
        all_issues.extend(data)

        if debug:
            print(f"🔹 Ambil batch: start={start}, count={len(data)}, total={len(all_issues)}")

        if len(data) < limit:
            break

        start += limit

    print(f"📦 Total issue keseluruhan: {len(all_issues)}")
    return all_issues


def fetch_issue_details(base_url, issue_id, auth, verify_tls, debug=False):
    """Ambil detail dari setiap issue"""
    headers = {
        "Authorization": f"{auth['token_type']} {auth['access_token']}",
        "Accept": "application/json"
    }
    url = f"{base_url}/api/v1/issueDetails/{issue_id}"
    resp = requests.get(url, headers=headers, verify=verify_tls)
    if resp.status_code != 200:
        if debug:
            print(f"⚠️  Gagal ambil detail issue {issue_id}: {resp.status_code}")
        return {}
    return resp.json().get("data", {})

# ============================================================
# Excel formatting
# ============================================================

def style_excel(file_path):
    """Tambahkan warna severity dan autofit kolom di Excel"""
    wb = load_workbook(file_path)
    ws = wb.active

    # Header bold
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Warnai baris berdasarkan severity
    severity_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value and str(cell.value).lower() == "severity":
            severity_col = idx
            break

    if severity_col:
        for row in ws.iter_rows(min_row=2):
            sev = row[severity_col - 1].value
            color = severity_color(sev)
            if color:
                for c in row:
                    c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Autofit column width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    wb.save(file_path)

# ============================================================
# Main export logic
# ============================================================

def fetch_and_export(base_url, username, password, project_version, output, verify_tls, engine_type=None, debug=False):
    """Ambil semua issue + detail lalu ekspor ke Excel"""
    if not verify_tls:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    if debug:
        logging.basicConfig(level=logging.DEBUG, format="%(levelname)s: %(message)s")

    auth = authenticate(base_url, username, password, verify_tls, debug)
    issues = fetch_issues(base_url, project_version, auth, verify_tls, debug)

    # Filter engine type jika bukan 'all'
    if engine_type and not (len(engine_type) == 1 and engine_type[0].lower() == "all"):
        engine_list = [e.lower() for e in engine_type]
        before = len(issues)
        issues = [i for i in issues if i.get("engineType", "").lower() in engine_list]
        print(f"🔍 Filter engineType={engine_list} → {before} ➜ {len(issues)} issues")
    else:
        print("⚙️  Mengambil semua engineType (no filter)")

    # Ambil detail dengan progress bar
    records = []
    for issue in tqdm(issues, desc="🔎 Fetching issue details", ncols=100):
        detail = fetch_issue_details(base_url, issue["id"], auth, verify_tls, debug)

        severity_value = (
            issue.get("friority")
            or issue.get("severityString")
            or issue.get("severity", "")
        )

        records.append({
            "Category": (issue.get("primaryTag") or {}).get("name", issue.get("issueName", "")),
            "Severity": severity_value,
            "Primary Location": issue.get("fullFileName", ""),
            "Url": detail.get("attackPayload", ""),
            "Line Number": issue.get("lineNumber", ""),
            "Overview": clean_html(detail.get("brief", "")),
            "Details": clean_html(detail.get("detail", "")),
            "Implication": clean_html(detail.get("tips", "")),
            "Recommendation": clean_html(detail.get("recommendation", "")),
            "Reference": clean_html(detail.get("references", "")),
            "Engine Type": issue.get("engineType", ""),
            "Found Date": issue.get("foundDate", ""),
            "Suppressed": issue.get("suppressed", ""),
            "Issue Status": issue.get("issueStatus", "")
        })

    df = pd.DataFrame(records)

    # Urutkan berdasarkan severity (kalau tersedia)
    if "Severity" in df.columns:
        df["Severity Rank"] = df["Severity"].apply(severity_rank)
        df.sort_values(by="Severity Rank", ascending=False, inplace=True)
        df.drop(columns=["Severity Rank"], inplace=True)

    # Urutan kolom sesuai template baru
    column_order = [
        "Category",
        "Severity",
        "Primary Location",
        "Line Number",
        "Url",
        "Overview",
        "Details",
        "Implication",
        "Recommendation",
        "Reference",
        "Engine Type",
        "Found Date",
        "Suppressed",
        "Issue Status"
    ]
    df = df[[col for col in column_order if col in df.columns]]

    # Simpan ke Excel dan format
    df.to_excel(output, index=False)
    style_excel(output)
    print(f"📊 File Excel berhasil dibuat dan diformat: {output}")

# ============================================================
# CLI Entry Point
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Export Fortify SSC issues ke Excel via API")
    parser.add_argument("--base-url", required=True)
    parser.add_argument("--user", required=True)
    parser.add_argument("--pass", dest="password", required=True)
    parser.add_argument("--project-version", required=True)
    parser.add_argument("--output", default="SSC_Issues_Report.xlsx")
    parser.add_argument("--insecure", action="store_true")
    parser.add_argument("--engine-type", nargs="+", help="Gunakan 'all' untuk ambil semua engine types")
    parser.add_argument("--debug", action="store_true")

    args = parser.parse_args()
    fetch_and_export(
        base_url=args.base_url,
        username=args.user,
        password=args.password,
        project_version=args.project_version,
        output=args.output,
        verify_tls=not args.insecure,
        engine_type=args.engine_type,
        debug=args.debug
    )

if __name__ == "__main__":
    main()
